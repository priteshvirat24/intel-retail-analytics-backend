"""
Generates complete reports and deliverables for the 52-Retailer Apify-Only Benchmark:
1. Markdown Report: reports/laptop_apify_52_final.md
2. JSON Dataset: reports/laptop_apify_52_final.json
3. Summary CSV: reports/laptop_apify_52_benchmark.csv
4. Excel Multi-Sheet Report: reports/laptop_apify_52_benchmark.xlsx
5. Side-by-Side Provider Comparison: reports/laptop_provider_comparison_52.md
"""
import os
import sys
import json
import csv
import statistics
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.models.registry import TargetRegistry

EVIDENCE_APIFY_BASE = PROJECT_ROOT / "evidence" / "apify"
EVIDENCE_BRD_BASE = PROJECT_ROOT / "evidence" / "brightdata"
REPORTS_DIR = PROJECT_ROOT / "reports"


def load_apify_results() -> Dict[str, Dict[str, Any]]:
    results = {}
    reg = TargetRegistry()
    for t in sorted(reg.all_targets(), key=lambda x: x.target_id):
        t_id = t.target_id
        ev_file = EVIDENCE_APIFY_BASE / t_id / "evidence_summary.json"
        if ev_file.exists():
            with open(ev_file, "r", encoding="utf-8") as f:
                results[t_id] = json.load(f)
        else:
            # Placeholder for unrun target
            results[t_id] = {
                "target_id": t_id,
                "retailer": t.retailer,
                "country": t.country,
                "domain": t.domain,
                "can_scrape": "NO",
                "status": "FAILURE",
                "access_success": False,
                "discovery_success": False,
                "extraction_success": False,
                "validation_success": False,
                "failure_stage": "ACCESS",
                "failure_category": "ACCESS_FAILURE",
                "failure_reason": "UNEXECUTED",
                "failure_message": "Target has not been crawled yet."
            }
    return results


def load_brightdata_results() -> Dict[str, Dict[str, Any]]:
    brd_results = {}
    reg = TargetRegistry()
    for t in sorted(reg.all_targets(), key=lambda x: x.target_id):
        t_id = t.target_id
        ev_file = EVIDENCE_BRD_BASE / t_id / "evidence_summary.json"
        if ev_file.exists():
            with open(ev_file, "r", encoding="utf-8") as f:
                brd_results[t_id] = json.load(f)
        else:
            brd_results[t_id] = {"can_scrape": "NO", "status": "FAILURE"}
    return brd_results


def compute_metrics(apify_data: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    total = len(apify_data)
    successful = sum(1 for r in apify_data.values() if r.get("status") == "SUCCESS" or r.get("can_scrape") == "YES")
    failed = total - successful
    success_rate = (successful / total * 100) if total > 0 else 0.0

    access_cnt = sum(1 for r in apify_data.values() if r.get("access_success"))
    disc_cnt = sum(1 for r in apify_data.values() if r.get("discovery_success"))
    ext_cnt = sum(1 for r in apify_data.values() if r.get("extraction_success"))
    val_cnt = sum(1 for r in apify_data.values() if r.get("validation_success"))

    durations = [r.get("execution_duration_sec", 0.0) for r in apify_data.values() if r.get("execution_duration_sec", 0) > 0]
    avg_dur = statistics.mean(durations) if durations else 0.0
    med_dur = statistics.median(durations) if durations else 0.0

    pages = [r.get("pages_crawled", 0) for r in apify_data.values()]
    avg_pages = statistics.mean(pages) if pages else 0.0
    succ_pages = [r.get("pages_crawled", 0) for r in apify_data.values() if r.get("status") == "SUCCESS"]
    avg_succ_pages = statistics.mean(succ_pages) if succ_pages else 0.0

    retries = [r.get("retry_count", 0) for r in apify_data.values()]
    avg_retries = statistics.mean(retries) if retries else 0.0

    waf = sum(1 for r in apify_data.values() if r.get("failure_category") in ("WAF_OR_ANTI_BOT", "ACCESS_FAILURE") or r.get("failure_reason") == "BOT_PROTECTION")
    disc_f = sum(1 for r in apify_data.values() if r.get("failure_category") == "URL_DISCOVERY_FAILURE" or r.get("failure_stage") == "DISCOVERY")
    ext_f = sum(1 for r in apify_data.values() if r.get("failure_category") == "EXTRACTION_FAILURE" or r.get("failure_stage") == "EXTRACTION")
    val_f = sum(1 for r in apify_data.values() if r.get("failure_category") == "VALIDATION_FAILURE" or r.get("failure_stage") == "VALIDATION")
    timeouts = sum(1 for r in apify_data.values() if r.get("failure_reason") in ("TIMEOUT", "APIFY_TIMEOUT"))
    actor_f = sum(1 for r in apify_data.values() if r.get("failure_reason") in ("APIFY_ACTOR_FAILURE", "APIFY_AUTH_FAILED"))

    return {
        "total_targets": total,
        "successful_targets": successful,
        "failed_targets": failed,
        "success_rate": f"{success_rate:.1f}%",
        "access_success_count": access_cnt,
        "discovery_success_count": disc_cnt,
        "extraction_success_count": ext_cnt,
        "validation_success_count": val_cnt,
        "average_execution_time_sec": round(avg_dur, 2),
        "median_execution_time_sec": round(med_dur, 2),
        "average_pages_crawled": round(avg_pages, 2),
        "average_pages_crawled_successful": round(avg_succ_pages, 2),
        "average_retries": round(avg_retries, 2),
        "waf_failures": waf,
        "discovery_failures": disc_f,
        "extraction_failures": ext_f,
        "validation_failures": val_f,
        "timeouts": timeouts,
        "actor_or_auth_failures": actor_f
    }


def generate_markdown_report(apify_data: Dict[str, Any], metrics: Dict[str, Any]) -> Path:
    out_file = REPORTS_DIR / "laptop_apify_52_final.md"
    lines = [
        "# 52-Retailer Laptop Crawling Benchmark — Apify-Only Independent Report",
        "",
        f"**Execution Date**: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"**Provider**: Apify Platform (`apify-client`)",
        f"**Benchmark Result**: **{metrics['successful_targets']} / {metrics['total_targets']} ({metrics['success_rate']})**",
        "",
        "## Executive Summary",
        "",
        f"- **Total Targets Evaluated**: {metrics['total_targets']}",
        f"- **Successful Targets**: {metrics['successful_targets']}",
        f"- **Failed Targets**: {metrics['failed_targets']}",
        f"- **Overall Success Rate**: {metrics['success_rate']}",
        "",
        "### Stage Funnel Progression",
        f"- **Page Access Success (`PAGE_FETCH_SUCCESS`)**: {metrics['access_success_count']} / {metrics['total_targets']}",
        f"- **Product Discovery Success (`PRODUCT_DISCOVERED`)**: {metrics['discovery_success_count']} / {metrics['total_targets']}",
        f"- **Product Extraction Success (`PRODUCT_EXTRACTED`)**: {metrics['extraction_success_count']} / {metrics['total_targets']}",
        f"- **Product Validation Success (`PRODUCT_VALIDATED`)**: {metrics['validation_success_count']} / {metrics['total_targets']}",
        "",
        "### Execution & Latency Telemetry",
        f"- **Average Duration**: {metrics['average_execution_time_sec']}s",
        f"- **Median Duration**: {metrics['median_execution_time_sec']}s",
        f"- **Average Pages Crawled per Target**: {metrics['average_pages_crawled']}",
        f"- **Average Pages for Successful Targets**: {metrics['average_pages_crawled_successful']}",
        f"- **Average Retries**: {metrics['average_retries']}",
        "",
        "### Failure Taxonomy Breakdown",
        f"- **WAF / Anti-Bot / Access Failures**: {metrics['waf_failures']}",
        f"- **URL / Product Discovery Failures**: {metrics['discovery_failures']}",
        f"- **Content Extraction Failures**: {metrics['extraction_failures']}",
        f"- **Classification / Validation Failures**: {metrics['validation_failures']}",
        f"- **Timeouts**: {metrics['timeouts']}",
        f"- **Actor / Authentication Failures**: {metrics['actor_or_auth_failures']}",
        "",
        "## Complete 52-Retailer Target Matrix",
        "",
        "| Target | Country | Access | Crawl | Product Found | SKU Extracted | Validated | Status | Failure Stage | Reason |",
        "| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |",
    ]

    for tid, r in sorted(apify_data.items()):
        cnt = r.get("country", "")
        acc = "✅" if r.get("access_success") else "❌"
        crw = "✅" if r.get("pages_crawled", 0) > 0 or r.get("discovery_success") else "❌"
        fnd = "✅" if r.get("discovery_success") else "❌"
        ext = "✅" if r.get("extraction_success") else "❌"
        val = "✅" if r.get("validation_success") else "❌"
        st = "✅ YES" if (r.get("status") == "SUCCESS" or r.get("can_scrape") == "YES") else "❌ NO"
        f_stage = r.get("failure_stage") or "NONE"
        f_reason = r.get("failure_reason") or "SUCCESS"
        lines.append(f"| `{tid}` | {cnt} | {acc} | {crw} | {fnd} | {ext} | {val} | {st} | `{f_stage}` | {f_reason} |")

    lines.extend([
        "",
        "## Evidence Files & Deliverables",
        "",
        "- **JSON Dataset**: [`laptop_apify_52_final.json`](file:///Users/priteshhome/crawl/reports/laptop_apify_52_final.json)",
        "- **Summary CSV**: [`laptop_apify_52_benchmark.csv`](file:///Users/priteshhome/crawl/reports/laptop_apify_52_benchmark.csv)",
        "- **Excel Report**: [`laptop_apify_52_benchmark.xlsx`](file:///Users/priteshhome/crawl/reports/laptop_apify_52_benchmark.xlsx)",
        "- **Side-by-Side Comparison**: [`laptop_provider_comparison_52.md`](file:///Users/priteshhome/crawl/reports/laptop_provider_comparison_52.md)",
        "- **Raw Evidence Directory**: `evidence/apify/<target_id>/`",
        "",
        "---",
        "*Report automatically generated by Apify Benchmark Suite.*"
    ])

    with open(out_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_file


def generate_json_report(apify_data: Dict[str, Any], metrics: Dict[str, Any]) -> Path:
    out_file = REPORTS_DIR / "laptop_apify_52_final.json"
    data = {
        "title": "52-Retailer Laptop Crawling Benchmark — Apify Independent Run",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "provider": "apify",
        "metrics": metrics,
        "results": apify_data
    }
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return out_file


def generate_csv_report(apify_data: Dict[str, Any]) -> Path:
    out_file = REPORTS_DIR / "laptop_apify_52_benchmark.csv"
    headers = [
        "target_id", "retailer", "country", "domain",
        "status", "can_scrape",
        "access_success", "discovery_success", "extraction_success", "validation_success",
        "failure_stage", "failure_category", "failure_reason", "failure_message",
        "strategy", "method", "actor_id", "actor_run_id",
        "execution_duration_sec", "pages_crawled", "retry_count",
        "title", "brand", "final_product_url"
    ]
    with open(out_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for tid, r in sorted(apify_data.items()):
            writer.writerow([
                r.get("target_id"),
                r.get("retailer"),
                r.get("country"),
                r.get("domain"),
                r.get("status"),
                r.get("can_scrape"),
                r.get("access_success"),
                r.get("discovery_success"),
                r.get("extraction_success"),
                r.get("validation_success"),
                r.get("failure_stage"),
                r.get("failure_category"),
                r.get("failure_reason"),
                r.get("failure_message"),
                r.get("strategy"),
                r.get("method"),
                r.get("actor_id"),
                r.get("actor_run_id"),
                r.get("execution_duration_sec"),
                r.get("pages_crawled"),
                r.get("retry_count"),
                r.get("title"),
                r.get("brand"),
                r.get("final_product_url")
            ])
    return out_file


def generate_comparison_report(apify_data: Dict[str, Any], brd_data: Dict[str, Any]) -> Path:
    out_file = REPORTS_DIR / "laptop_provider_comparison_52.md"
    
    both_success = 0
    brd_only = 0
    apify_only = 0
    both_fail = 0

    comp_rows = []
    for tid in sorted(set(list(apify_data.keys()) + list(brd_data.keys()))):
        a_res = apify_data.get(tid, {})
        b_res = brd_data.get(tid, {})
        
        a_ok = (a_res.get("status") == "SUCCESS" or a_res.get("can_scrape") == "YES")
        b_ok = (b_res.get("status") == "SUCCESS" or b_res.get("can_scrape") == "YES")
        
        if a_ok and b_ok:
            result_tag = "BOTH"
            both_success += 1
        elif b_ok and not a_ok:
            result_tag = "BRIGHT_DATA_ONLY"
            brd_only += 1
        elif a_ok and not b_ok:
            result_tag = "APIFY_ONLY"
            apify_only += 1
        else:
            result_tag = "BOTH_FAIL"
            both_fail += 1

        comp_rows.append({
            "target_id": tid,
            "retailer": a_res.get("retailer") or b_res.get("retailer") or tid,
            "country": a_res.get("country") or b_res.get("country") or "",
            "brightdata": "SUCCESS" if b_ok else "FAILURE",
            "apify": "SUCCESS" if a_ok else "FAILURE",
            "result": result_tag,
            "apify_reason": a_res.get("failure_reason") or "SUCCESS",
            "brd_method": b_res.get("strategy") or "BRIGHTDATA_WEB_UNLOCKER"
        })

    brd_success_total = both_success + brd_only
    apify_success_total = both_success + apify_only

    lines = [
        "# 52-Retailer Scraping Benchmark: Bright Data vs. Apify Comparison",
        "",
        f"**Execution Date**: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "## Executive Summary Matrix",
        "",
        "| Metric | Bright Data | Apify | Comparison / Delta |",
        "| :--- | :---: | :---: | :---: |",
        f"| **Total Targets** | 52 | 52 | Same 52 Canonical Targets |",
        f"| **Successful Crawls** | **{brd_success_total}** ({brd_success_total/52*100:.1f}%) | **{apify_success_total}** ({apify_success_total/52*100:.1f}%) | {brd_success_total - apify_success_total:+d} for Bright Data |",
        f"| **Both Succeeded** | {both_success} | {both_success} | Overlap Targets |",
        f"| **Bright Data Only** | {brd_only} | - | Bright Data Superiority |",
        f"| **Apify Only** | - | {apify_only} | Apify Superiority |",
        f"| **Both Failed** | {both_fail} | {both_fail} | Hard Anti-Bot Targets |",
        "",
        "---",
        "",
        "## Detailed Question Analysis",
        "",
        "### A. Can Apify crawl the website?",
        f"- Apify achieved page access across **{sum(1 for r in apify_data.values() if r.get('access_success'))} / 52** targets.",
        f"- Heavily protected sites with advanced WAFs (Akamai, Cloudflare Turnstile, Kasada, PerimeterX) require residential unblocking proxies.",
        "",
        "### B. Can Apify reach a relevant page?",
        f"- Category seed and search query discovery succeeded on **{sum(1 for r in apify_data.values() if r.get('discovery_success'))} / 52** targets.",
        "",
        "### C. Can Apify discover the actual product page?",
        f"- Candidates matching laptop URL patterns were identified for **{sum(1 for r in apify_data.values() if r.get('discovery_success'))} / 52** targets.",
        "",
        "### D. Can Apify extract the product/SKU?",
        f"- Product metadata and title extraction completed on **{sum(1 for r in apify_data.values() if r.get('extraction_success'))} / 52** targets.",
        "",
        "### E. Can we validate the extracted result?",
        f"- Strict classification via `LaptopClassifier` confirmed genuine laptops on **{apify_success_total} / 52** targets.",
        "",
        "### F. If it fails, exactly why?",
        "- Primary failure stages documented in the taxonomy breakdown (e.g. `ACCESS_FAILURE` / `WAF_OR_ANTI_BOT` / `APIFY_AUTH_FAILED` / `URL_DISCOVERY_FAILURE`).",
        "",
        "---",
        "",
        "## Per-Target Side-by-Side Comparison Table",
        "",
        "| # | Target ID | Retailer Name | Country | Bright Data | Apify | Outcome | Apify Reason |",
        "|---|-----------|---------------|---------|:-----------:|:-----:|:-------:|--------------|",
    ]

    for idx, row in enumerate(comp_rows, 1):
        b_icon = "✅ SUCCESS" if row["brightdata"] == "SUCCESS" else "❌ FAILURE"
        a_icon = "✅ SUCCESS" if row["apify"] == "SUCCESS" else "❌ FAILURE"
        tag = f"`{row['result']}`"
        lines.append(f"| {idx} | `{row['target_id']}` | {row['retailer']} | {row['country']} | {b_icon} | {a_icon} | {tag} | {row['apify_reason']} |")

    lines.extend([
        "",
        "---",
        "*Benchmark comparisons generated from independent empirical execution logs.*"
    ])

    with open(out_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_file


def generate_excel_report(apify_data: Dict[str, Any], brd_data: Dict[str, Any], metrics: Dict[str, Any]) -> Path:
    out_file = REPORTS_DIR / "laptop_apify_52_benchmark.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    FONT_FAMILY = "Segoe UI"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    
    zebra_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    zebra_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    success_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    success_font = Font(name=FONT_FAMILY, size=10, bold=True, color="166534")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fail_font = Font(name=FONT_FAMILY, size=10, bold=True, color="991B1B")

    data_font = Font(name=FONT_FAMILY, size=10, color="334155")
    bold_font = Font(name=FONT_FAMILY, size=10, bold=True, color="1E293B")
    link_font = Font(name=FONT_FAMILY, size=9, color="2563EB", underline="single")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # -------------------------------------------------------------
    # SHEET 1: 52 Retailers Apify Benchmark
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="52 Retailers Apify Benchmark")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:K1")
    t1 = ws1["A1"]
    t1.value = "52-Retailer Apify-Only Independent Laptop Crawling Benchmark"
    t1.font = Font(name=FONT_FAMILY, size=14, bold=True, color="FFFFFF")
    t1.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    t1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:K2")
    sub1 = ws1["A2"]
    sub1.value = f"Success Rate: {metrics['successful_targets']}/52 ({metrics['success_rate']}) | Execution Date: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | Provider: Apify Platform"
    sub1.font = Font(name=FONT_FAMILY, size=9, italic=True, color="94A3B8")
    sub1.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 20

    ws1_headers = [
        ("#", 6), ("Target ID", 18), ("Retailer", 20), ("Country", 16),
        ("Access", 10), ("Crawl", 10), ("Product Found", 14), ("SKU Extracted", 14),
        ("Validated", 12), ("Status", 12), ("Failure Reason / Details", 35)
    ]

    for col_idx, (h_name, width) in enumerate(ws1_headers, 1):
        cell = ws1.cell(row=4, column=col_idx)
        cell.value = h_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5, 6, 7, 8, 9, 10] else "left", vertical="center")
        cell.border = thin_border
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = width
    ws1.row_dimensions[4].height = 26

    r_num = 5
    for idx, (tid, r) in enumerate(sorted(apify_data.items()), 1):
        is_even = (idx % 2 == 0)
        row_fill = zebra_even if is_even else zebra_odd
        is_ok = (r.get("status") == "SUCCESS" or r.get("can_scrape") == "YES")

        vals = [
            idx,
            tid,
            r.get("retailer", tid),
            r.get("country", ""),
            "YES" if r.get("access_success") else "NO",
            "YES" if r.get("pages_crawled", 0) > 0 or r.get("discovery_success") else "NO",
            "YES" if r.get("discovery_success") else "NO",
            "YES" if r.get("extraction_success") else "NO",
            "YES" if r.get("validation_success") else "NO",
            "SUCCESS" if is_ok else "FAILURE",
            r.get("failure_reason") or "SUCCESS"
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=r_num, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = data_font
            
            if col_idx in [1, 5, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.font = bold_font
            elif col_idx == 10:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = success_fill if is_ok else fail_fill
                cell.font = success_font if is_ok else fail_font

        ws1.row_dimensions[r_num].height = 20
        r_num += 1
    ws1.freeze_panes = "A5"

    # -------------------------------------------------------------
    # SHEET 2: BrightData vs Apify Comparison
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="BrightData vs Apify Comparison")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value = "Apples-to-Apples Provider Benchmark: Bright Data vs. Apify (52 Targets)"
    t2.font = Font(name=FONT_FAMILY, size=13, bold=True, color="FFFFFF")
    t2.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 32

    ws2_headers = [
        ("#", 6), ("Target ID", 18), ("Retailer", 20), ("Country", 16),
        ("Bright Data", 16), ("Apify", 16), ("Comparison Result", 22), ("Apify Diagnostic Reason", 32)
    ]

    for col_idx, (h_name, width) in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.value = h_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5, 6, 7] else "left", vertical="center")
        cell.border = thin_border
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = width
    ws2.row_dimensions[3].height = 26

    r2_num = 4
    for idx, tid in enumerate(sorted(set(list(apify_data.keys()) + list(brd_data.keys()))), 1):
        is_even = (idx % 2 == 0)
        row_fill = zebra_even if is_even else zebra_odd

        a_res = apify_data.get(tid, {})
        b_res = brd_data.get(tid, {})
        a_ok = (a_res.get("status") == "SUCCESS" or a_res.get("can_scrape") == "YES")
        b_ok = (b_res.get("status") == "SUCCESS" or b_res.get("can_scrape") == "YES")

        if a_ok and b_ok:
            tag = "BOTH"
        elif b_ok and not a_ok:
            tag = "BRIGHT_DATA_ONLY"
        elif a_ok and not b_ok:
            tag = "APIFY_ONLY"
        else:
            tag = "BOTH_FAIL"

        vals2 = [
            idx,
            tid,
            a_res.get("retailer") or b_res.get("retailer") or tid,
            a_res.get("country") or b_res.get("country") or "",
            "SUCCESS" if b_ok else "FAILURE",
            "SUCCESS" if a_ok else "FAILURE",
            tag,
            a_res.get("failure_reason") or "SUCCESS"
        ]

        for col_idx, val in enumerate(vals2, 1):
            cell = ws2.cell(row=r2_num, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = data_font

            if col_idx in [1, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.font = bold_font
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = success_fill if b_ok else fail_fill
                cell.font = success_font if b_ok else fail_font
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = success_fill if a_ok else fail_fill
                cell.font = success_font if a_ok else fail_font

        ws2.row_dimensions[r2_num].height = 20
        r2_num += 1
    ws2.freeze_panes = "A4"

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_file))
    return out_file


def main():
    print("Generating comprehensive Apify benchmark reports and comparison deliverables...")
    apify_data = load_apify_results()
    brd_data = load_brightdata_results()
    metrics = compute_metrics(apify_data)

    md_path = generate_markdown_report(apify_data, metrics)
    json_path = generate_json_report(apify_data, metrics)
    csv_path = generate_csv_report(apify_data)
    comp_path = generate_comparison_report(apify_data, brd_data)
    xlsx_path = generate_excel_report(apify_data, brd_data, metrics)

    print(f"✅ Generated Markdown:   {md_path}")
    print(f"✅ Generated JSON:       {json_path}")
    print(f"✅ Generated CSV:        {csv_path}")
    print(f"✅ Generated Comparison: {comp_path}")
    print(f"✅ Generated Excel:      {xlsx_path}")


if __name__ == "__main__":
    main()
