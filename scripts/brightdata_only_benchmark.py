"""
Dedicated Bright Data ONLY Laptop Scraping Benchmark & Analytics Generator.
Uses official Bright Data Async Web Unlocker API (POST /unblocker/req & GET /unblocker/get_result)
across all 52 canonical targets.

Outputs to NEW dedicated deliverables (without modifying previous files):
- reports/brightdata_only_52_site_scrape_analytics.csv
- reports/brightdata_only_laptop_benchmark.xlsx
"""
import os
import re
import csv
import json
import time
import asyncio
from pathlib import Path
from typing import Dict, Any, List, Optional
import httpx
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.registry import TargetRegistry
from app.crawlers.base import CrawlerResponse
from app.evaluation.laptop_validator import LaptopValidator


BRIGHTDATA_API_KEY = os.getenv("BRIGHTDATA_API_KEY", "")
BRIGHTDATA_ZONE = os.getenv("BRIGHTDATA_WEB_UNLOCKER_ZONE", os.getenv("BRIGHTDATA_ZONE", "web_unlocker1"))
BRIGHTDATA_CUSTOMER = os.getenv("BRIGHTDATA_CUSTOMER_ID", "")

REQ_URL = f"https://api.brightdata.com/unblocker/req?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"
GET_RESULT_BASE = f"https://api.brightdata.com/unblocker/get_result?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"


async def fetch_brightdata_unlocker(client: httpx.AsyncClient, url: str, timeout: float = 30.0) -> CrawlerResponse:
    """Fetches a page through Bright Data Async Web Unlocker REST API."""
    if not url or url == "NONE" or not url.startswith("http"):
        return CrawlerResponse(
            url=url or "NONE",
            final_url=url or "NONE",
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="NO_LAPTOP_URL_DISCOVERED",
            error_message="No product URL available to crawl",
            response_time_ms=0.0
        )

    headers = {
        "Authorization": f"Bearer {BRIGHTDATA_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {"url": url}
    t0 = time.perf_counter()

    try:
        r = await client.post(REQ_URL, headers=headers, json=payload, timeout=12.0)
        response_id = r.headers.get("x-response-id")

        if not response_id or r.status_code not in (200, 202):
            lat_ms = (time.perf_counter() - t0) * 1000.0
            err_msg = r.headers.get("x-brd-err-msg") or r.text[:200]
            return CrawlerResponse(
                url=url,
                final_url=url,
                status_code=r.status_code,
                strategy="BRIGHTDATA_WEB_UNLOCKER",
                success=False,
                error_message=err_msg,
                failure_reason=r.headers.get("x-brd-err-code", f"HTTP_{r.status_code}"),
                response_time_ms=round(lat_ms, 1)
            )

        # Poll for async result
        poll_url = f"{GET_RESULT_BASE}&response_id={response_id}"
        max_attempts = int(timeout / 2.0)

        for _ in range(max_attempts):
            await asyncio.sleep(2.0)
            res_r = await client.get(poll_url, headers=headers, timeout=12.0)
            if res_r.status_code == 200:
                lat_ms = (time.perf_counter() - t0) * 1000.0
                return CrawlerResponse(
                    url=url,
                    final_url=url,
                    status_code=200,
                    html=res_r.text,
                    headers=dict(res_r.headers),
                    strategy="BRIGHTDATA_WEB_UNLOCKER",
                    success=True,
                    response_time_ms=round(lat_ms, 1)
                )
            elif res_r.status_code == 202:
                # Still processing
                continue
            else:
                lat_ms = (time.perf_counter() - t0) * 1000.0
                return CrawlerResponse(
                    url=url,
                    final_url=url,
                    status_code=res_r.status_code,
                    strategy="BRIGHTDATA_WEB_UNLOCKER",
                    success=False,
                    failure_reason=f"HTTP_{res_r.status_code}",
                    error_message=res_r.text[:200],
                    response_time_ms=round(lat_ms, 1)
                )

        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TIMEOUT",
            error_message="Unlocker async polling timed out after 30s",
            response_time_ms=round(lat_ms, 1)
        )

    except Exception as e:
        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TRANSPORT_FAILURE",
            error_message=str(e),
            response_time_ms=round(lat_ms, 1)
        )


async def run_brightdata_benchmark():
    registry = TargetRegistry("config/targets.yaml")
    all_targets = registry.all_targets()
    print(f"=== Starting Bright Data ONLY Scraping Benchmark across {len(all_targets)} Retailers ===", flush=True)

    # Load previously discovered frozen laptop URLs
    prev_json = Path("reports/laptop_crawl_benchmark.json")
    seed_map = {}
    if prev_json.exists():
        with open(prev_json, "r", encoding="utf-8") as f:
            d = json.load(f)
            for row in d.get("matrix", []):
                ret = row.get("retailer", "").strip().lower()
                cnt = row.get("country", "").strip().lower()
                url = row.get("laptop_url")
                if url and url != "NONE":
                    seed_map[(ret, cnt)] = url

    results: List[Dict[str, Any]] = []
    sem = asyncio.Semaphore(4)

    async def _process_target(idx: int, target):
        async with sem:
            retailer_name = target.brand_name
            country = target.country
            target_id = target.target_id

            # Candidate URL
            existing_url = seed_map.get((retailer_name.lower(), country.lower()))
            if existing_url and existing_url != "NONE":
                laptop_url = existing_url
            else:
                seed_urls = []
                if getattr(target, "discovery_seeds", None):
                    seed_urls.extend(target.discovery_seeds)
                if getattr(target, "sample_product_urls", None):
                    seed_urls.extend(target.sample_product_urls)
                if getattr(target, "category_urls", None):
                    seed_urls.extend(target.category_urls)
                laptop_url = seed_urls[0] if seed_urls else "NONE"

            async with httpx.AsyncClient() as client:
                print(f"[{idx:02d}/{len(all_targets)}] Scraping {retailer_name} ({country}) -> {laptop_url[:60]}...", flush=True)
                resp = await fetch_brightdata_unlocker(client, laptop_url, timeout=30.0)

                # Validate & Extract Product
                val = LaptopValidator.validate(resp, laptop_url, threshold=0.75)
                can_scrape = "YES" if val.is_valid_laptop else "NO"

                # Extract Product Details
                prod_title = val.product_name or "—"
                brand = val.brand or "—"
                price_str = f"{val.price} {val.currency or 'USD'}" if val.price else "—"
                sku = val.model_or_sku or "—"

                if can_scrape == "YES":
                    reason = "CRAWL_SUCCESS: Authentic laptop product page verified & extracted via Bright Data"
                else:
                    if laptop_url == "NONE":
                        reason = "Store accessible but no laptop product link could be identified"
                    elif resp.headers.get("x-brd-err-msg"):
                        reason = f"Bright Data Zone Error: {resp.headers.get('x-brd-err-msg')}"
                    elif resp.status_code == 0:
                        reason = f"Bright Data Timeout: {resp.error_message}"
                    elif resp.status_code in (403, 429):
                        reason = f"Anti-bot blocked with HTTP {resp.status_code} ({val.failure_vendor or 'WAF'})"
                    elif resp.status_code == 404:
                        reason = "Regional geoblock returned HTTP 404 Not Found"
                    elif "EMPTY" in (val.failure_class or ""):
                        reason = "Client-side SPA shell returned empty HTML without product schema"
                    elif "LOW_CONFIDENCE" in (val.failure_class or ""):
                        reason = "Category/guide page reached instead of single laptop SKU"
                    elif "NOT_A_LAPTOP" in (val.failure_class or ""):
                        reason = "Non-laptop accessory product discovered"
                    else:
                        reason = val.failure_class or f"Extraction incomplete (Confidence: {val.confidence_score})"

                print(f"  -> [{idx:02d}/{len(all_targets)}] {retailer_name} ({country}): Can Scrape = {can_scrape} (Status: {resp.status_code}, Score: {val.confidence_score})", flush=True)

                evidence_folder = f"evidence/{target.retailer.lower().replace(' ', '_')}/{target.country.lower().replace(' ', '_')}/laptop/brightdata/"

                # Save Evidence
                ev_dir = Path(evidence_folder)
                ev_dir.mkdir(parents=True, exist_ok=True)
                with open(ev_dir / "crawl_attempt.json", "w", encoding="utf-8") as f:
                    json.dump({
                        "target_id": target_id,
                        "retailer": retailer_name,
                        "country": country,
                        "url": laptop_url,
                        "status_code": resp.status_code,
                        "response_time_ms": resp.response_time_ms,
                        "can_scrape": can_scrape,
                        "confidence_score": val.confidence_score,
                        "product_name": val.product_name,
                        "brand": val.brand,
                        "price": val.price,
                        "currency": val.currency,
                        "sku": val.model_or_sku,
                        "reason": reason
                    }, f, indent=2)

                if resp.html:
                    with open(ev_dir / "raw.html", "w", encoding="utf-8", errors="ignore") as f:
                        f.write(resp.html)

                res_dict = {
                    "#": idx,
                    "Retailer Name": retailer_name,
                    "Country / Region": country,
                    "Can Scrape Laptop Data?": can_scrape,
                    "Scraped Laptop Product Title": prod_title if can_scrape == "YES" else "—",
                    "Brand": brand if can_scrape == "YES" else "—",
                    "Price & Currency": price_str if can_scrape == "YES" else "—",
                    "Model / SKU": sku if can_scrape == "YES" else "—",
                    "Tested Product Page URL": laptop_url,
                    "Reason If Cannot Scrape (Failure Root Cause)": reason,
                    "Strategy Used": "Bright Data Web Unlocker",
                    "Forensic Evidence Folder": evidence_folder,
                    "status_code": resp.status_code,
                    "response_time_ms": resp.response_time_ms
                }
                return res_dict

    tasks = [_process_target(i, t) for i, t in enumerate(all_targets, 1)]
    results = await asyncio.gather(*tasks)

    # Sort results by #
    results.sort(key=lambda x: x["#"])

    # 1. GENERATE NEW DEDICATED CSV
    out_csv = Path("reports/brightdata_only_52_site_scrape_analytics.csv")
    csv_headers = [
        "#",
        "Retailer Name",
        "Country / Region",
        "Can Scrape Laptop Data?",
        "Scraped Laptop Product Title",
        "Brand",
        "Price & Currency",
        "Model / SKU",
        "Tested Product Page URL",
        "Reason If Cannot Scrape (Failure Root Cause)",
        "Strategy Used",
        "Forensic Evidence Folder"
    ]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    print(f"\n[SUCCESS] Dedicated Bright Data CSV generated: {out_csv}", flush=True)

    # 2. GENERATE NEW DEDICATED EXCEL
    out_xlsx = Path("reports/brightdata_only_laptop_benchmark.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Bright Data 52-Site Analytics")
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_regular = Font(name="Calibri", size=11, color="1E293B")

    fill_teal_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_success = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_failed = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border_side = Side(style="thin", color="CBD5E1")
    border_card = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "52-Target Bright Data ONLY Laptop Scraping Benchmark"
    ws["A1"].font = font_title
    ws["A2"] = f"Standalone audit executed via Bright Data Web Unlocker API (Zone: web_unlocker1) | Generated: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())}"
    ws["A2"].font = font_subtitle

    for c_idx, h in enumerate(csv_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_teal_header
        cell.alignment = align_center
        cell.border = border_card

    for r_idx, row_dict in enumerate(results, start=5):
        can_scrape = row_dict["Can Scrape Laptop Data?"]
        for c_idx, h in enumerate(csv_headers, start=1):
            val = row_dict[h]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_card

            if c_idx == 1:
                cell.alignment = align_center
                cell.font = font_bold
            elif c_idx in [2, 3]:
                cell.alignment = align_left
                cell.font = font_bold
            elif c_idx == 4:
                cell.alignment = align_center
                cell.font = font_bold
                cell.fill = fill_success if can_scrape == "YES" else fill_failed
            elif c_idx in [5, 6, 7, 8]:
                cell.alignment = align_left if c_idx == 5 else align_center
                if can_scrape == "YES":
                    cell.font = font_bold
                    cell.fill = fill_highlight
            elif c_idx in [9, 10]:
                cell.alignment = align_left
                if can_scrape == "YES" and c_idx == 10:
                    cell.fill = fill_success
                    cell.font = font_bold
            elif c_idx in [11, 12]:
                cell.alignment = align_center if c_idx == 11 else align_left

    # Adjust column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 36
    ws.column_dimensions["J"].width = 45
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 35

    wb.save(out_xlsx)
    print(f"[SUCCESS] Dedicated Bright Data Excel generated: {out_xlsx}", flush=True)

    succ_count = sum(1 for r in results if r["Can Scrape Laptop Data?"] == "YES")
    print(f"\n=== Benchmark Complete: {succ_count} / {len(results)} Targets Successfully Scraped via Bright Data ({round(succ_count/len(results)*100, 1)}%) ===", flush=True)


if __name__ == "__main__":
    asyncio.run(run_brightdata_benchmark())
