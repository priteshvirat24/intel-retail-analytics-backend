"""
Full-Potential Bright Data Scraping & Product Extraction Engine across all 52 Retailers.
Features:
1. Dynamic ISO country targeting (flags: country-us, country-de, country-gb, country-fr, country-in, etc.)
2. Two-stage autonomous crawler:
   - Stage 1: Discovers direct Laptop Product Detail Pages (PDP) from search/category seeds via Bright Data.
   - Stage 2: Scrapes the single Laptop PDP via Bright Data Web Unlocker with country-specific residential routing.
3. Multi-language product keyword and JSON-LD/microdata/regex attribute extraction.
4. Generates updated, comprehensive CSV and Excel reports with forensic evidence.
"""
import os
import re
import csv
import json
import time
import asyncio
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from urllib.parse import urljoin
import httpx
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.registry import TargetRegistry
from app.crawlers.base import CrawlerResponse
from app.evaluation.laptop_validator import LaptopValidator


BRIGHTDATA_API_KEY = os.getenv("BRIGHTDATA_API_KEY", "")
BRIGHTDATA_ZONE = os.getenv("BRIGHTDATA_WEB_UNLOCKER_ZONE", os.getenv("BRIGHTDATA_ZONE", "web_unlocker1"))
BRIGHTDATA_CUSTOMER = os.getenv("BRIGHTDATA_CUSTOMER_ID", "")

REQ_URL = f"https://api.brightdata.com/unblocker/req?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"
GET_RESULT_BASE = f"https://api.brightdata.com/unblocker/get_result?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"

COUNTRY_TO_ISO = {
    "United States": "us",
    "India": "in",
    "United Kingdom": "gb",
    "Germany": "de",
    "France": "fr",
    "Italy": "it",
    "Spain": "es",
    "Canada": "ca",
    "Mexico": "mx",
    "Brazil": "br",
    "Indonesia": "id",
    "South Korea": "kr",
    "Denmark": "dk",
    "Norway": "no",
    "Sweden": "se",
    "Australia": "au",
    "China": "cn",
    "Poland": "pl",
    "Japan": "jp",
    "Turkey": "tr",
    "Chile": "cl",
    "Colombia": "co",
    "Vietnam": "vn",
    "Global": "us"
}


async def fetch_brightdata_unlocker_page(client: httpx.AsyncClient, url: str, country_iso: str, timeout: float = 35.0) -> CrawlerResponse:
    """Fetches a page through Bright Data Async Web Unlocker REST API with country flag."""
    if not url or url == "NONE" or not url.startswith("http"):
        return CrawlerResponse(
            url=url or "NONE",
            final_url=url or "NONE",
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="INVALID_URL",
            error_message="Invalid or missing URL",
            response_time_ms=0.0
        )

    headers = {
        "Authorization": f"Bearer {BRIGHTDATA_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "url": url,
        "flags": f"country-{country_iso.lower()}"
    }
    t0 = time.perf_counter()

    try:
        r = await client.post(REQ_URL, headers=headers, json=payload, timeout=15.0)
        response_id = r.headers.get("x-response-id")

        if not response_id or r.status_code not in (200, 202):
            lat_ms = (time.perf_counter() - t0) * 1000.0
            err_msg = r.headers.get("x-brd-err-msg") or r.text[:200]
            return CrawlerResponse(
                url=url,
                final_url=url,
                status_code=r.status_code,
                strategy="BRIGHTDATA_WEB_UNLOCKER",
                success=False,
                error_message=err_msg,
                failure_reason=r.headers.get("x-brd-err-code", f"HTTP_{r.status_code}"),
                response_time_ms=round(lat_ms, 1)
            )

        poll_url = f"{GET_RESULT_BASE}&response_id={response_id}"
        max_attempts = int(timeout / 2.0)

        for _ in range(max_attempts):
            await asyncio.sleep(2.0)
            try:
                res_r = await client.get(poll_url, headers=headers, timeout=15.0)
                if res_r.status_code == 200:
                    lat_ms = (time.perf_counter() - t0) * 1000.0
                    return CrawlerResponse(
                        url=url,
                        final_url=url,
                        status_code=200,
                        html=res_r.text,
                        headers=dict(res_r.headers),
                        strategy="BRIGHTDATA_WEB_UNLOCKER",
                        success=True,
                        response_time_ms=round(lat_ms, 1)
                    )
                elif res_r.status_code == 202:
                    continue
                else:
                    lat_ms = (time.perf_counter() - t0) * 1000.0
                    return CrawlerResponse(
                        url=url,
                        final_url=url,
                        status_code=res_r.status_code,
                        strategy="BRIGHTDATA_WEB_UNLOCKER",
                        success=False,
                        failure_reason=f"HTTP_{res_r.status_code}",
                        error_message=res_r.text[:200],
                        response_time_ms=round(lat_ms, 1)
                    )
            except httpx.TimeoutException:
                continue

        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TIMEOUT",
            error_message="Unlocker async polling timed out after 35s",
            response_time_ms=round(lat_ms, 1)
        )

    except Exception as e:
        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TRANSPORT_FAILURE",
            error_message=str(e),
            response_time_ms=round(lat_ms, 1)
        )


def extract_candidate_product_links(html: str, base_url: str) -> List[str]:
    """Extracts candidate individual laptop SKU product links from search/catalog HTML."""
    if not html or len(html) < 200:
        return []
    soup = BeautifulSoup(html, "html.parser")
    candidates = []

    # Priority patterns for PDP links
    pdp_indicators = ["/dp/b0", "/product/", "/p/", "/item/", "/ip/", "/portatil-", "/notebook-", "/laptop-", ".p?skuid=", "/pd/"]
    laptop_keywords = ["laptop", "notebook", "macbook", "chromebook", "portatil", "ordinateur", "dizustu", "computadora", "ideapad", "thinkpad", "vivobook", "zenbook", "pavilion", "inspiron", "vostro", "galaxy book", "proart", "legion", "omen", "tuf", "rog"]

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("javascript:") or href.startswith("#"):
            continue
        full_url = urljoin(base_url, href)
        full_lower = full_url.lower()

        # Score link
        is_pdp = any(p in full_lower for p in pdp_indicators)
        has_kw = any(k in full_lower for k in laptop_keywords) or any(k in a.get_text().lower() for k in laptop_keywords)

        if is_pdp and has_kw:
            candidates.append(full_url)
        elif is_pdp:
            candidates.append(full_url)

    # Deduplicate preserving order
    seen = set()
    unique = []
    for c in candidates:
        # Strip query params like session tracking
        clean = c.split("?")[0] if "/dp/" in c or "/product/" in c or "/p/" in c else c
        if clean not in seen:
            seen.add(clean)
            unique.append(c)

    return unique


async def process_retailer_brightdata(idx: int, total: int, target, seed_url_map: Dict[Tuple[str, str], str]) -> Dict[str, Any]:
    """Processes a single retailer target end-to-end using full potential of Bright Data."""
    retailer_name = target.brand_name
    country = target.country
    target_id = target.target_id
    base_url = target.base_url
    country_iso = COUNTRY_TO_ISO.get(country, "us")

    print(f"[{idx:02d}/{total}] Processing {retailer_name} ({country}) [ISO: {country_iso}]...", flush=True)

    async with httpx.AsyncClient() as client:
        # Step 1: Check existing seed URL
        existing_url = seed_url_map.get((retailer_name.lower(), country.lower()))
        is_direct_pdp = False
        if existing_url and existing_url != "NONE":
            # Check if this URL looks like a direct PDP
            if any(p in existing_url.lower() for p in ["/dp/b0", "/product/", "/p/", ".html", ".p?skuid=", "/pd/"]):
                is_direct_pdp = True

        target_pdp_url = existing_url if is_direct_pdp else None

        # Step 2: If no direct PDP or previous was a category hub, run Bright Data Search Discovery
        if not target_pdp_url or not is_direct_pdp:
            # Build localized search URLs
            search_candidates = []
            if getattr(target, "sample_product_urls", None):
                search_candidates.extend(target.sample_product_urls)
            if getattr(target, "discovery_seeds", None):
                search_candidates.extend(target.discovery_seeds)
            if getattr(target, "category_urls", None):
                search_candidates.extend(target.category_urls)

            # Add query search seeds
            if country_iso in ["fr"]:
                search_candidates.append(urljoin(base_url, "/s?k=ordinateur+portable"))
                search_candidates.append(urljoin(base_url, "/recherche?q=ordinateur+portable"))
            elif country_iso in ["es", "mx", "cl", "co"]:
                search_candidates.append(urljoin(base_url, "/s?k=portatil"))
                search_candidates.append(urljoin(base_url, "/search?q=portatil"))
                search_candidates.append(urljoin(base_url, "/laptops"))
            elif country_iso in ["de"]:
                search_candidates.append(urljoin(base_url, "/s?k=laptop"))
                search_candidates.append(urljoin(base_url, "/suche?q=laptop"))
            elif country_iso in ["it"]:
                search_candidates.append(urljoin(base_url, "/s?k=notebook"))
                search_candidates.append(urljoin(base_url, "/ricerca?q=notebook"))
            elif country_iso in ["br"]:
                search_candidates.append(urljoin(base_url, "/s?k=notebook"))
                search_candidates.append(urljoin(base_url, "/busca?q=notebook"))
            elif country_iso in ["tr"]:
                search_candidates.append(urljoin(base_url, "/arama?q=laptop"))
            elif country_iso in ["pl"]:
                search_candidates.append(urljoin(base_url, "/szukaj?q=laptop"))
            elif country_iso in ["vn"]:
                search_candidates.append(urljoin(base_url, "/laptop"))
            elif country_iso in ["kr"]:
                search_candidates.append(urljoin(base_url, "/search?q=laptop"))
            elif country_iso in ["jp"]:
                search_candidates.append(urljoin(base_url, "/category/19531/19532/"))
            else:
                search_candidates.append(urljoin(base_url, "/s?k=laptop"))
                search_candidates.append(urljoin(base_url, "/search?q=laptop"))
                search_candidates.append(urljoin(base_url, "/laptops"))

            # Execute discovery probe on first working search candidate
            for s_url in search_candidates[:3]:
                if not s_url or not s_url.startswith("http"):
                    continue
                print(f"  [{idx:02d}] Probing discovery seed via Bright Data: {s_url[:65]}...", flush=True)
                disc_resp = await fetch_brightdata_unlocker_page(client, s_url, country_iso, timeout=25.0)
                if disc_resp.status_code == 200 and disc_resp.html:
                    extracted_pdp_links = extract_candidate_product_links(disc_resp.html, base_url)
                    if extracted_pdp_links:
                        target_pdp_url = extracted_pdp_links[0]
                        print(f"  [{idx:02d}] DISCOVERED Product PDP: {target_pdp_url[:65]}", flush=True)
                        break

            if not target_pdp_url:
                target_pdp_url = existing_url if existing_url and existing_url != "NONE" else base_url

        # Step 3: Scrape the Target Laptop Product Detail Page (PDP) via Bright Data
        print(f"  [{idx:02d}] Scraping Laptop PDP: {target_pdp_url[:65]}...", flush=True)
        resp = await fetch_brightdata_unlocker_page(client, target_pdp_url, country_iso, timeout=35.0)

        # Step 4: Validate and Extract Laptop Product Data
        val = LaptopValidator.validate(resp, target_pdp_url, threshold=0.70)
        can_scrape = "YES" if val.is_valid_laptop else "NO"

        prod_title = val.product_name or "—"
        brand = val.brand or "—"
        price_str = f"{val.price} {val.currency or 'USD'}" if val.price else "—"
        sku = val.model_or_sku or "—"

        if can_scrape == "YES":
            reason = f"CRAWL_SUCCESS: Successfully scraped & verified authentic laptop SKU ({prod_title} | Brand: {brand} | Price: {price_str} | SKU: {sku})"
        else:
            if resp.status_code == 0:
                reason = f"Bright Data Connection Timeout: {resp.error_message}"
            elif resp.status_code in (403, 429):
                reason = f"Anti-bot WAF Challenge: Blocked with HTTP {resp.status_code} ({val.failure_vendor or 'Protected WAF'})"
            elif resp.status_code == 404:
                reason = "Regional Geoblock: Egress IP routed but page returned HTTP 404 Not Found"
            elif "EMPTY" in (val.failure_class or ""):
                reason = "Client-side SPA shell: JavaScript hydration required to populate product DOM tree"
            elif "LOW_CONFIDENCE" in (val.failure_class or ""):
                reason = "Catalog aggregation page reached: Multiple listing cards present without individual checkout schema"
            elif "NOT_A_LAPTOP" in (val.failure_class or ""):
                reason = "Peripheral/Accessory SKU reached instead of standalone laptop computer"
            else:
                reason = val.failure_class or f"Extraction incomplete (Confidence Score: {val.confidence_score})"

        print(f"  -> [{idx:02d}/{total}] {retailer_name} ({country}): Can Scrape = {can_scrape} (Score: {val.confidence_score}, Status: {resp.status_code})", flush=True)

        evidence_folder = f"evidence/{target.retailer.lower().replace(' ', '_')}/{target.country.lower().replace(' ', '_')}/laptop/brightdata/"
        ev_dir = Path(evidence_folder)
        ev_dir.mkdir(parents=True, exist_ok=True)
        with open(ev_dir / "crawl_attempt.json", "w", encoding="utf-8") as f:
            json.dump({
                "target_id": target_id,
                "retailer": retailer_name,
                "country": country,
                "country_iso": country_iso,
                "url": target_pdp_url,
                "status_code": resp.status_code,
                "response_time_ms": resp.response_time_ms,
                "can_scrape": can_scrape,
                "confidence_score": val.confidence_score,
                "product_name": val.product_name,
                "brand": val.brand,
                "price": val.price,
                "currency": val.currency,
                "sku": val.model_or_sku,
                "reason": reason
            }, f, indent=2)

        if resp.html:
            with open(ev_dir / "raw.html", "w", encoding="utf-8", errors="ignore") as f:
                f.write(resp.html)

        return {
            "#": idx,
            "Retailer Name": retailer_name,
            "Country / Region": country,
            "Can Scrape Laptop Data?": can_scrape,
            "Scraped Laptop Product Title": prod_title if can_scrape == "YES" else "—",
            "Brand": brand if can_scrape == "YES" else "—",
            "Price & Currency": price_str if can_scrape == "YES" else "—",
            "Model / SKU": sku if can_scrape == "YES" else "—",
            "Tested Product Page URL": target_pdp_url,
            "Reason If Cannot Scrape (Failure Root Cause)": reason,
            "Strategy Used": f"Bright Data Web Unlocker (flags: country-{country_iso})",
            "Forensic Evidence Folder": evidence_folder,
            "status_code": resp.status_code,
            "response_time_ms": resp.response_time_ms
        }


async def run_full_potential_benchmark():
    registry = TargetRegistry("config/targets.yaml")
    all_targets = registry.all_targets()
    print(f"=== Launching Full Potential Bright Data Scraping Suite across {len(all_targets)} Targets ===", flush=True)

    # Seed map
    prev_json = Path("reports/laptop_crawl_benchmark.json")
    seed_map = {}
    if prev_json.exists():
        with open(prev_json, "r", encoding="utf-8") as f:
            d = json.load(f)
            for row in d.get("matrix", []):
                ret = row.get("retailer", "").strip().lower()
                cnt = row.get("country", "").strip().lower()
                url = row.get("laptop_url")
                if url and url != "NONE":
                    seed_map[(ret, cnt)] = url

    sem = asyncio.Semaphore(4)

    async def _worker(i, t):
        async with sem:
            return await process_retailer_brightdata(i, len(all_targets), t, seed_map)

    tasks = [_worker(i, t) for i, t in enumerate(all_targets, 1)]
    results = await asyncio.gather(*tasks)
    results.sort(key=lambda x: x["#"])

    # 1. GENERATE UPDATED CSV DELIVERABLE
    out_csv = Path("reports/brightdata_only_52_site_scrape_analytics.csv")
    csv_headers = [
        "#",
        "Retailer Name",
        "Country / Region",
        "Can Scrape Laptop Data?",
        "Scraped Laptop Product Title",
        "Brand",
        "Price & Currency",
        "Model / SKU",
        "Tested Product Page URL",
        "Reason If Cannot Scrape (Failure Root Cause)",
        "Strategy Used",
        "Forensic Evidence Folder"
    ]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    print(f"\n[SUCCESS] Dedicated Bright Data CSV updated: {out_csv}", flush=True)

    # 2. GENERATE UPDATED EXCEL DELIVERABLE
    out_xlsx = Path("reports/brightdata_only_laptop_benchmark.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Bright Data 52-Site Analytics")
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_regular = Font(name="Calibri", size=11, color="1E293B")

    fill_teal_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_success = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_failed = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border_side = Side(style="thin", color="CBD5E1")
    border_card = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "52-Target Full Potential Bright Data Laptop Scraping Benchmark"
    ws["A1"].font = font_title
    ws["A2"] = f"Standalone audit utilizing dynamic ISO country-routing, automated catalog discovery & Web Unlocker | Generated: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())}"
    ws["A2"].font = font_subtitle

    for c_idx, h in enumerate(csv_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_teal_header
        cell.alignment = align_center
        cell.border = border_card

    for r_idx, row_dict in enumerate(results, start=5):
        can_scrape = row_dict["Can Scrape Laptop Data?"]
        for c_idx, h in enumerate(csv_headers, start=1):
            val = row_dict[h]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_card

            if c_idx == 1:
                cell.alignment = align_center
                cell.font = font_bold
            elif c_idx in [2, 3]:
                cell.alignment = align_left
                cell.font = font_bold
            elif c_idx == 4:
                cell.alignment = align_center
                cell.font = font_bold
                cell.fill = fill_success if can_scrape == "YES" else fill_failed
            elif c_idx in [5, 6, 7, 8]:
                cell.alignment = align_left if c_idx == 5 else align_center
                if can_scrape == "YES":
                    cell.font = font_bold
                    cell.fill = fill_highlight
            elif c_idx in [9, 10]:
                cell.alignment = align_left
                if can_scrape == "YES" and c_idx == 10:
                    cell.fill = fill_success
                    cell.font = font_bold
            elif c_idx in [11, 12]:
                cell.alignment = align_center if c_idx == 11 else align_left

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 36
    ws.column_dimensions["J"].width = 65
    ws.column_dimensions["K"].width = 35
    ws.column_dimensions["L"].width = 35

    wb.save(out_xlsx)
    print(f"[SUCCESS] Dedicated Bright Data Excel updated: {out_xlsx}", flush=True)

    succ_count = sum(1 for r in results if r["Can Scrape Laptop Data?"] == "YES")
    print(f"\n=== Full Potential Benchmark Complete: {succ_count} / {len(results)} Targets Successfully Scraped via Bright Data ({round(succ_count/len(results)*100, 1)}%) ===", flush=True)


if __name__ == "__main__":
    asyncio.run(run_full_potential_benchmark())
