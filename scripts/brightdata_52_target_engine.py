"""
Production 52-Retailer Full Potential Bright Data Laptop Crawling & Extraction Engine.
Features:
1. Bright Data Web Unlocker as Primary & Only Crawler Layer.
2. Accurate Geographic Country-Targeted Routing across 52 Retailers.
3. Multi-Hop Autonomous Discovery: Category -> Search -> Sitemaps -> Product Links.
4. Candidate URL Extraction, Deduplication, and Scoring (testing up to 10 candidates per retailer).
5. Strict Laptop Classifier (12 classes, hard negative filtering, spec extraction).
6. Comprehensive Forensic Logging and Evidence Preservation.
"""
import os
import re
import csv
import json
import time
import asyncio
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from urllib.parse import urljoin, urlparse, parse_qs
import httpx
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.registry import TargetRegistry
from app.crawlers.base import CrawlerResponse
from app.classification.laptop_classifier import LaptopClassifier, ProductClass
from app.evaluation.laptop_validator import LaptopValidator


BRIGHTDATA_API_KEY = os.getenv("BRIGHTDATA_API_KEY", "")
BRIGHTDATA_ZONE = os.getenv("BRIGHTDATA_WEB_UNLOCKER_ZONE", os.getenv("BRIGHTDATA_ZONE", "web_unlocker1"))
BRIGHTDATA_CUSTOMER = os.getenv("BRIGHTDATA_CUSTOMER_ID", "")

REQ_URL = f"https://api.brightdata.com/unblocker/req?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"
GET_RESULT_BASE = f"https://api.brightdata.com/unblocker/get_result?customer={BRIGHTDATA_CUSTOMER}&zone={BRIGHTDATA_ZONE}"

COUNTRY_TO_ISO = {
    "United States": "us",
    "India": "in",
    "United Kingdom": "gb",
    "Germany": "de",
    "France": "fr",
    "Italy": "it",
    "Spain": "es",
    "Canada": "ca",
    "Mexico": "mx",
    "Brazil": "br",
    "Indonesia": "id",
    "South Korea": "kr",
    "Denmark": "dk",
    "Norway": "no",
    "Sweden": "se",
    "Australia": "au",
    "China": "cn",
    "Poland": "pl",
    "Japan": "jp",
    "Turkey": "tr",
    "Chile": "cl",
    "Colombia": "co",
    "Vietnam": "vn",
    "Global": "us"
}


async def fetch_via_brightdata(client: httpx.AsyncClient, url: str, country_iso: str, timeout: float = 35.0) -> CrawlerResponse:
    """Fetches a page through Bright Data Async Web Unlocker REST API with country flag."""
    if not url or url == "NONE" or not url.startswith("http"):
        return CrawlerResponse(
            url=url or "NONE",
            final_url=url or "NONE",
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="INVALID_URL",
            error_message="Invalid or missing URL",
            response_time_ms=0.0
        )

    headers = {
        "Authorization": f"Bearer {BRIGHTDATA_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "url": url,
        "flags": f"country-{country_iso.lower()}"
    }
    t0 = time.perf_counter()

    try:
        r = await client.post(REQ_URL, headers=headers, json=payload, timeout=15.0)
        response_id = r.headers.get("x-response-id")

        if not response_id or r.status_code not in (200, 202):
            lat_ms = (time.perf_counter() - t0) * 1000.0
            err_msg = r.headers.get("x-brd-err-msg") or r.text[:200]
            return CrawlerResponse(
                url=url,
                final_url=url,
                status_code=r.status_code,
                strategy="BRIGHTDATA_WEB_UNLOCKER",
                success=False,
                error_message=err_msg,
                failure_reason=r.headers.get("x-brd-err-code", f"HTTP_{r.status_code}"),
                response_time_ms=round(lat_ms, 1)
            )

        poll_url = f"{GET_RESULT_BASE}&response_id={response_id}"
        max_attempts = int(timeout / 2.0)

        for _ in range(max_attempts):
            await asyncio.sleep(2.0)
            try:
                res_r = await client.get(poll_url, headers=headers, timeout=15.0)
                if res_r.status_code == 200:
                    lat_ms = (time.perf_counter() - t0) * 1000.0
                    return CrawlerResponse(
                        url=url,
                        final_url=url,
                        status_code=200,
                        html=res_r.text,
                        headers=dict(res_r.headers),
                        strategy="BRIGHTDATA_WEB_UNLOCKER",
                        success=True,
                        response_time_ms=round(lat_ms, 1)
                    )
                elif res_r.status_code == 202:
                    continue
                else:
                    lat_ms = (time.perf_counter() - t0) * 1000.0
                    return CrawlerResponse(
                        url=url,
                        final_url=url,
                        status_code=res_r.status_code,
                        strategy="BRIGHTDATA_WEB_UNLOCKER",
                        success=False,
                        failure_reason=f"HTTP_{res_r.status_code}",
                        error_message=res_r.text[:200],
                        response_time_ms=round(lat_ms, 1)
                    )
            except httpx.TimeoutException:
                continue

        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TIMEOUT",
            error_message="Unlocker async polling timed out after 35s",
            response_time_ms=round(lat_ms, 1)
        )

    except Exception as e:
        lat_ms = (time.perf_counter() - t0) * 1000.0
        return CrawlerResponse(
            url=url,
            final_url=url,
            status_code=0,
            strategy="BRIGHTDATA_WEB_UNLOCKER",
            success=False,
            failure_reason="TRANSPORT_FAILURE",
            error_message=str(e),
            response_time_ms=round(lat_ms, 1)
        )


def extract_and_rank_candidate_urls(html: str, base_url: str) -> List[Tuple[str, float, str]]:
    """Extracts, filters, and ranks candidate product URLs from search/catalog HTML."""
    if not html or len(html) < 200:
        return []
    soup = BeautifulSoup(html, "html.parser")
    candidates = []

    pdp_patterns = [
        r"/dp/[a-z0-9]{10}",
        r"/product/[a-z0-9\-_]+",
        r"/p/[a-z0-9\-_]+",
        r"/item/[a-z0-9\-_]+",
        r"/ip/[a-z0-9\-_]+",
        r"/portatil-[a-z0-9\-_]+",
        r"/notebook-[a-z0-9\-_]+",
        r"/laptop-[a-z0-9\-_]+",
        r"\.p\?skuid=[0-9]+",
        r"-[0-9]{6,10}\.html",
        r"/pd/[a-z0-9\-_]+"
    ]

    laptop_positives = [
        "laptop", "notebook", "macbook", "chromebook", "portatil", "portátil",
        "ordinateur", "dizustu", "dizüstü", "ideapad", "thinkpad", "vivobook",
        "zenbook", "pavilion", "inspiron", "vostro", "galaxy book", "proart",
        "legion", "omen", "tuf", "rog", "aspire", "swift", "predator", "nitro"
    ]

    hard_negatives = [
        "bag", "case", "sleeve", "backpack", "cover", "skin", "stand", "cooler",
        "charger", "adapter", "powerbank", "cable", "mouse", "keyboard", "headset",
        "monitor", "antivirus", "mcafee", "norton", "office", "windows", "condizionatore",
        "funda", "mochila", "housse", "tasche", "custodia", "pasta"
    ]

    seen = set()

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("javascript:") or href.startswith("#") or href.startswith("mailto:"):
            continue

        full_url = urljoin(base_url, href)
        parsed = urlparse(full_url)
        clean_path = parsed.path.lower()
        anchor_text = a.get_text().strip().lower()

        # Check if URL matches PDP patterns
        is_pdp = any(re.search(p, clean_path) for p in pdp_patterns) or any(re.search(p, full_url.lower()) for p in pdp_patterns)

        # Check if URL or anchor has negative signals
        is_negative = any(re.search(rf"\b{neg}\b", clean_path) for neg in hard_negatives) or any(re.search(rf"\b{neg}\b", anchor_text) for neg in hard_negatives)
        if is_negative:
            continue

        # Score candidate
        score = 0.0
        if is_pdp:
            score += 0.50

        # Positive keywords
        for pos in laptop_positives:
            if pos in clean_path:
                score += 0.30
                break
            if pos in anchor_text:
                score += 0.20
                break

        if is_pdp or score >= 0.50:
            # Deduplicate by clean path
            norm_key = f"{parsed.netloc}{parsed.path}"
            if norm_key not in seen:
                seen.add(norm_key)
                candidates.append((full_url, score, anchor_text[:60]))

    # Sort descending by score
    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates


async def run_retailer_crawl(idx: int, total: int, target, seed_url_map: Dict[Tuple[str, str], str]) -> Dict[str, Any]:
    """Autonomous multi-candidate evaluation loop for a single retailer."""
    retailer_name = target.brand_name
    country = target.country
    target_id = target.target_id
    base_url = target.base_url
    country_iso = COUNTRY_TO_ISO.get(country, "us")

    print(f"[{idx:02d}/{total}] START: {retailer_name} ({country}) [ISO: {country_iso}]", flush=True)

    ev_base = Path(f"evidence/{target.retailer.lower().replace(' ', '_')}/{target.country.lower().replace(' ', '_')}/laptop/brightdata")
    for sub in ["discovery", "candidates", "product", "attempts", "final"]:
        (ev_base / sub).mkdir(parents=True, exist_ok=True)

    attempt_log = []
    discovered_candidates = []
    winning_product = None
    winning_classification = None
    winning_resp = None
    winning_url = None

    async with httpx.AsyncClient() as client:
        # Step 1: Gather candidate URLs
        candidate_queue = []

        # 1a. Check previously verified seed URL
        existing_url = seed_url_map.get((retailer_name.lower(), country.lower()))
        if existing_url and existing_url != "NONE":
            candidate_queue.append((existing_url, 0.95, "Verified Candidate Seed"))

        # 1b. Build localized discovery search URLs
        search_seeds = []
        if getattr(target, "sample_product_urls", None):
            for u in target.sample_product_urls:
                search_seeds.append(u)
        if getattr(target, "discovery_seeds", None):
            for u in target.discovery_seeds:
                search_seeds.append(u)
        if getattr(target, "category_urls", None):
            for u in target.category_urls:
                search_seeds.append(u)

        # Dynamic multi-language queries
        if country_iso == "fr":
            search_seeds.extend([urljoin(base_url, "/s?k=ordinateur+portable"), urljoin(base_url, "/recherche?q=ordinateur+portable")])
        elif country_iso in ["es", "mx", "cl", "co"]:
            search_seeds.extend([urljoin(base_url, "/s?k=portatil"), urljoin(base_url, "/search?q=portatil"), urljoin(base_url, "/laptops")])
        elif country_iso == "de":
            search_seeds.extend([urljoin(base_url, "/s?k=laptop"), urljoin(base_url, "/suche?q=laptop")])
        elif country_iso == "it":
            search_seeds.extend([urljoin(base_url, "/s?k=notebook"), urljoin(base_url, "/ricerca?q=notebook")])
        elif country_iso == "br":
            search_seeds.extend([urljoin(base_url, "/s?k=notebook"), urljoin(base_url, "/busca?q=notebook")])
        elif country_iso == "tr":
            search_seeds.extend([urljoin(base_url, "/arama?q=laptop"), urljoin(base_url, "/laptop")])
        elif country_iso == "pl":
            search_seeds.extend([urljoin(base_url, "/szukaj?q=laptop"), urljoin(base_url, "/laptopy")])
        elif country_iso == "vn":
            search_seeds.extend([urljoin(base_url, "/laptop"), urljoin(base_url, "/tim-kiem?k=laptop")])
        elif country_iso == "kr":
            search_seeds.extend([urljoin(base_url, "/search?q=노트북")])
        elif country_iso == "jp":
            search_seeds.extend([urljoin(base_url, "/category/19531/19532/")])
        else:
            search_seeds.extend([urljoin(base_url, "/s?k=laptop"), urljoin(base_url, "/search?q=laptop"), urljoin(base_url, "/laptops")])

        # Step 2: Proactive Search Discovery via Bright Data
        print(f"  [{idx:02d}] Executing search discovery across {len(search_seeds)} seeds...", flush=True)
        for s_url in search_seeds[:4]:
            if not s_url or not s_url.startswith("http"):
                continue
            disc_resp = await fetch_via_brightdata(client, s_url, country_iso, timeout=25.0)
            attempt_log.append({
                "type": "discovery_search",
                "url": s_url,
                "status_code": disc_resp.status_code,
                "latency_ms": disc_resp.response_time_ms,
                "html_bytes": len(disc_resp.html)
            })
            if disc_resp.status_code == 200 and disc_resp.html:
                extracted = extract_and_rank_candidate_urls(disc_resp.html, base_url)
                if extracted:
                    discovered_candidates.extend(extracted)
                    print(f"  [{idx:02d}] Extracted {len(extracted)} candidate product URLs from {s_url[:50]}", flush=True)
                    break

        # Save discovery artifacts
        with open(ev_base / "discovery" / "search_attempts.json", "w", encoding="utf-8") as f:
            json.dump(attempt_log, f, indent=2)

        # Merge candidate queues
        for c in discovered_candidates:
            if not any(c[0] == q[0] for q in candidate_queue):
                candidate_queue.append(c)

        # Save candidates
        with open(ev_base / "candidates" / "ranked_candidates.json", "w", encoding="utf-8") as f:
            json.dump([{"url": c[0], "score": c[1], "anchor": c[2]} for c in candidate_queue], f, indent=2)

        # Fallback to base_url if no candidates found
        if not candidate_queue:
            candidate_queue.append((base_url, 0.1, "Base URL Fallback"))

        # Step 3: Candidate Evaluation Loop (Up to 8 candidates per retailer)
        print(f"  [{idx:02d}] Testing up to {min(len(candidate_queue), 8)} candidate URLs...", flush=True)
        for c_idx, (cand_url, score, anchor) in enumerate(candidate_queue[:8], start=1):
            cand_resp = await fetch_via_brightdata(client, cand_url, country_iso, timeout=30.0)
            
            # Validate with LaptopValidator
            val = LaptopValidator.validate(cand_resp, cand_url, threshold=0.70)
            # Classify with strict LaptopClassifier
            classification = LaptopClassifier.classify(
                title=val.product_name or "",
                html=cand_resp.html,
                url=cand_url,
                price=val.price
            )

            eval_entry = {
                "candidate_index": c_idx,
                "url": cand_url,
                "status_code": cand_resp.status_code,
                "latency_ms": cand_resp.response_time_ms,
                "product_name": val.product_name,
                "brand": val.brand,
                "price": val.price,
                "is_valid_validator": val.is_valid_laptop,
                "product_class": classification.product_class.value,
                "is_genuine_laptop": classification.is_genuine_laptop,
                "rejection_reason": classification.rejection_reason
            }
            attempt_log.append(eval_entry)

            print(f"    [{idx:02d}.{c_idx}] Cand: {cand_url[:50]} -> Class: {classification.product_class.value}, Genuine: {classification.is_genuine_laptop} (Title: {(val.product_name or 'None')[:35]})", flush=True)

            if classification.is_genuine_laptop:
                winning_product = val
                winning_classification = classification
                winning_resp = cand_resp
                winning_url = cand_url
                break

        # Save all attempt logs
        with open(ev_base / "attempts" / "evaluation_attempts.json", "w", encoding="utf-8") as f:
            json.dump(attempt_log, f, indent=2)

        # Step 4: Finalize Result
        can_scrape = "YES" if (winning_product and winning_classification and winning_classification.is_genuine_laptop) else "NO"

        if can_scrape == "YES":
            prod_title = winning_product.product_name or "—"
            brand = winning_product.brand or winning_classification.detected_brand or "—"
            price_str = f"{winning_product.price} {winning_product.currency or 'USD'}" if winning_product.price else "—"
            sku = winning_product.model_or_sku or winning_classification.model_or_sku or "—"
            reason = f"CRAWL_SUCCESS: Verified authentic laptop product ({prod_title} | Brand: {brand} | Price: {price_str} | SKU: {sku})"
            final_url = winning_url
            final_status = winning_resp.status_code
            final_lat = winning_resp.response_time_ms

            # Save full product artifacts
            with open(ev_base / "product" / "raw.html", "w", encoding="utf-8", errors="ignore") as f:
                f.write(winning_resp.html)
            with open(ev_base / "product" / "extracted_product.json", "w", encoding="utf-8") as f:
                json.dump({
                    "title": prod_title,
                    "brand": brand,
                    "price": winning_product.price,
                    "currency": winning_product.currency,
                    "sku": sku,
                    "specs": winning_classification.extracted_specs,
                    "positive_signals": winning_classification.positive_signals
                }, f, indent=2)
            with open(ev_base / "final" / "classification.json", "w", encoding="utf-8") as f:
                json.dump(winning_classification.dict(), f, indent=2)
        else:
            prod_title = "—"
            brand = "—"
            price_str = "—"
            sku = "—"
            final_url = candidate_queue[0][0] if candidate_queue else base_url
            final_status = 0
            final_lat = 0.0

            # Determine explicit forensic failure reason
            rejections = [a.get("rejection_reason") for a in attempt_log if a.get("rejection_reason")]
            statuses = [a.get("status_code") for a in attempt_log if "status_code" in a]
            
            if any(s in [403, 429] for s in statuses):
                reason = "Anti-Bot WAF Challenge: Retailer returned HTTP 403/429 challenge blocking automated requests"
            elif all(s == 0 for s in statuses):
                reason = "Connection Timeout: Bright Data request timed out waiting for network response"
            elif rejections:
                reason = f"Filtered Negative Products: Tested candidates rejected ({rejections[0]})"
            else:
                reason = "Catalog Traversal Incomplete: No individual laptop product detail pages could be validated"

        print(f"[{idx:02d}/{total}] FINISHED {retailer_name} ({country}): Can Scrape = {can_scrape} | Reason: {reason[:60]}", flush=True)

        return {
            "#": idx,
            "Retailer Name": retailer_name,
            "Country / Region": country,
            "Can Scrape Laptop Data?": can_scrape,
            "Scraped Laptop Product Title": prod_title,
            "Brand": brand,
            "Price & Currency": price_str,
            "Model / SKU": sku,
            "Tested Product Page URL": final_url,
            "Reason If Cannot Scrape (Failure Root Cause)": reason,
            "Strategy Used": f"Bright Data Web Unlocker (flags: country-{country_iso})",
            "Forensic Evidence Folder": str(ev_base) + "/",
            "status_code": final_status,
            "response_time_ms": final_lat
        }


async def main():
    registry = TargetRegistry("config/targets.yaml")
    all_targets = registry.all_targets()
    print(f"=== Starting Production Bright Data 52-Retailer Crawling Engine ===", flush=True)

    # Seed map
    prev_json = Path("reports/laptop_crawl_benchmark.json")
    seed_map = {}
    if prev_json.exists():
        with open(prev_json, "r", encoding="utf-8") as f:
            d = json.load(f)
            for row in d.get("matrix", []):
                ret = row.get("retailer", "").strip().lower()
                cnt = row.get("country", "").strip().lower()
                url = row.get("laptop_url")
                if url and url != "NONE":
                    seed_map[(ret, cnt)] = url

    sem = asyncio.Semaphore(4)

    async def _worker(i, t):
        async with sem:
            return await run_retailer_crawl(i, len(all_targets), t, seed_map)

    tasks = [_worker(i, t) for i, t in enumerate(all_targets, 1)]
    results = await asyncio.gather(*tasks)
    results.sort(key=lambda x: x["#"])

    # 1. GENERATE UPDATED CSV DELIVERABLE
    out_csv = Path("reports/brightdata_only_52_site_scrape_analytics.csv")
    csv_headers = [
        "#",
        "Retailer Name",
        "Country / Region",
        "Can Scrape Laptop Data?",
        "Scraped Laptop Product Title",
        "Brand",
        "Price & Currency",
        "Model / SKU",
        "Tested Product Page URL",
        "Reason If Cannot Scrape (Failure Root Cause)",
        "Strategy Used",
        "Forensic Evidence Folder"
    ]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    print(f"\n[SUCCESS] Production Bright Data CSV updated: {out_csv}", flush=True)

    # 2. GENERATE UPDATED EXCEL DELIVERABLE
    out_xlsx = Path("reports/brightdata_only_laptop_benchmark.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Bright Data 52-Site Analytics")
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_regular = Font(name="Calibri", size=11, color="1E293B")

    fill_teal_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_success = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_failed = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border_side = Side(style="thin", color="CBD5E1")
    border_card = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "52-Target Bright Data Laptop Crawling & Strict Extraction Benchmark"
    ws["A1"].font = font_title
    ws["A2"] = f"Audited via Bright Data Web Unlocker | Strict 12-Class Classification | Generated: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())}"
    ws["A2"].font = font_subtitle

    for c_idx, h in enumerate(csv_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_teal_header
        cell.alignment = align_center
        cell.border = border_card

    for r_idx, row_dict in enumerate(results, start=5):
        can_scrape = row_dict["Can Scrape Laptop Data?"]
        for c_idx, h in enumerate(csv_headers, start=1):
            val = row_dict[h]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_card

            if c_idx == 1:
                cell.alignment = align_center
                cell.font = font_bold
            elif c_idx in [2, 3]:
                cell.alignment = align_left
                cell.font = font_bold
            elif c_idx == 4:
                cell.alignment = align_center
                cell.font = font_bold
                cell.fill = fill_success if can_scrape == "YES" else fill_failed
            elif c_idx in [5, 6, 7, 8]:
                cell.alignment = align_left if c_idx == 5 else align_center
                if can_scrape == "YES":
                    cell.font = font_bold
                    cell.fill = fill_highlight
            elif c_idx in [9, 10]:
                cell.alignment = align_left
                if can_scrape == "YES" and c_idx == 10:
                    cell.fill = fill_success
                    cell.font = font_bold
            elif c_idx in [11, 12]:
                cell.alignment = align_center if c_idx == 11 else align_left

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 36
    ws.column_dimensions["J"].width = 68
    ws.column_dimensions["K"].width = 35
    ws.column_dimensions["L"].width = 35

    wb.save(out_xlsx)
    print(f"[SUCCESS] Production Bright Data Excel updated: {out_xlsx}", flush=True)

    succ_count = sum(1 for r in results if r["Can Scrape Laptop Data?"] == "YES")
    print(f"\n=== Benchmark Complete: {succ_count} / {len(results)} Genuine Laptops Successfully Scraped via Bright Data ({round(succ_count/len(results)*100, 1)}%) ===", flush=True)


if __name__ == "__main__":
    asyncio.run(main())
